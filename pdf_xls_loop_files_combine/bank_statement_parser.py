#!/usr/bin/env python3
"""
Bank of America Statement Parser
Extracts transactions from PDF statements and exports to Excel
"""

import pdfplumber
import re
import os
import glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


class BankStatementParser:
    def __init__(self, pdf_path, output_dir):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.transactions = []
        
    def extract_text_from_pdf(self):
        """Extract all text from the PDF file"""
        text_content = []
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_content.append(text)
            return '\n'.join(text_content)
        except Exception as e:
            print(f"Error reading PDF: {e}")
            return None
    
    def parse_transactions(self, text):
        """Parse transactions from extracted text"""
        transactions = []
        
        # Patterns to identify transaction sections
        deposit_pattern = r'Deposits and other credits'
        withdrawal_pattern = r'Withdrawals and other debits'
        withdrawal_continued_pattern = r'Withdrawals and other debits\s*-\s*continued'
        
        lines = text.split('\n')
        
        # Find section markers
        in_deposits = False
        in_withdrawals = False
        in_withdrawals_continued = False
        
        for i, line in enumerate(lines):
            # Check for section headers
            if re.search(deposit_pattern, line, re.IGNORECASE):
                in_deposits = True
                in_withdrawals = False
                in_withdrawals_continued = False
                continue
            elif re.search(withdrawal_continued_pattern, line, re.IGNORECASE):
                in_withdrawals_continued = True
                in_deposits = False
                in_withdrawals = False
                continue
            elif re.search(withdrawal_pattern, line, re.IGNORECASE):
                in_withdrawals = True
                in_deposits = False
                in_withdrawals_continued = False
                continue
            
            # Check if we've moved to a new section (like Summary, Balance, etc.)
            if re.search(r'^\s*(Summary|Balance|Account|Statement|Page)', line, re.IGNORECASE):
                if not any([in_deposits, in_withdrawals, in_withdrawals_continued]):
                    continue
                # Check if this is actually a transaction line or a section header
                if not self._looks_like_transaction(line):
                    in_deposits = False
                    in_withdrawals = False
                    in_withdrawals_continued = False
                    continue
            
            # Parse transaction lines
            if in_deposits or in_withdrawals or in_withdrawals_continued:
                transaction = self._parse_transaction_line(line, in_deposits)
                if transaction:
                    transactions.append(transaction)
        
        return transactions
    
    def _looks_like_transaction(self, line):
        """Check if a line looks like a transaction"""
        # Transactions typically have dates and amounts
        date_pattern = r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}'
        amount_pattern = r'[\d,]+\.\d{2}'
        
        has_date = bool(re.search(date_pattern, line))
        has_amount = bool(re.search(amount_pattern, line))
        
        return has_date and has_amount
    
    def _parse_transaction_line(self, line, is_deposit):
        """Parse a single transaction line"""
        # Skip empty lines or lines that don't look like transactions
        if not line.strip() or len(line.strip()) < 10:
            return None
        
        # Pattern for date (MM/DD/YYYY or MM/DD/YY)
        date_pattern = r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
        # Pattern for amount (with commas and decimals)
        amount_pattern = r'([\d,]+\.\d{2})'
        
        # Find date
        date_match = re.search(date_pattern, line)
        if not date_match:
            return None
        
        date_str = date_match.group(1)
        
        # Find amount (usually at the end of the line)
        amount_matches = re.findall(amount_pattern, line)
        if not amount_matches:
            return None
        
        # The last number is usually the balance, second to last is the transaction amount
        # Or the first amount after the date might be the transaction amount
        amount_str = amount_matches[-1] if len(amount_matches) == 1 else amount_matches[-2]
        amount = float(amount_str.replace(',', ''))
        
        # Extract description (everything between date and amount)
        date_end = date_match.end()
        amount_start = line.rfind(amount_str)
        
        description = line[date_end:amount_start].strip()
        # Clean up description
        description = re.sub(r'\s+', ' ', description)
        
        # Determine transaction type
        transaction_type = 'Deposit' if is_deposit else 'Withdrawal'
        
        # Parse date
        try:
            if '/' in date_str:
                date_parts = date_str.split('/')
            else:
                date_parts = date_str.split('-')
            
            month = int(date_parts[0])
            day = int(date_parts[1])
            year = int(date_parts[2])
            if year < 100:
                year += 2000
            
            parsed_date = datetime(year, month, day)
        except:
            parsed_date = date_str
        
        return {
            'Date': parsed_date.strftime('%Y-%m-%d') if isinstance(parsed_date, datetime) else date_str,
            'Type': transaction_type,
            'Description': description,
            'Amount': amount if is_deposit else -amount,  # Negative for withdrawals
            'Raw_Line': line.strip()
        }
    
    def extract_table_transactions(self, text):
        """Alternative method: Try to extract transactions using table detection"""
        transactions = []
        
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # Try to extract tables
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if not table:
                            continue
                        
                        # Look for transaction tables
                        for row in table:
                            if not row or len(row) < 3:
                                continue
                            
                            # Check if row contains transaction data
                            row_text = ' '.join([str(cell) if cell else '' for cell in row])
                            
                            # Check for date pattern
                            if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', row_text):
                                transaction = self._parse_table_row(row, row_text)
                                if transaction:
                                    transactions.append(transaction)
        except Exception as e:
            print(f"Error extracting tables: {e}")
        
        return transactions
    
    def _parse_table_row(self, row, row_text):
        """Parse a transaction from a table row"""
        # Find date
        date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', row_text)
        if not date_match:
            return None
        
        date_str = date_match.group(1)
        
        # Find amounts
        amount_pattern = r'([\d,]+\.\d{2})'
        amounts = re.findall(amount_pattern, row_text)
        if not amounts:
            return None
        
        # Determine if deposit or withdrawal based on context
        is_deposit = 'deposit' in row_text.lower() or 'credit' in row_text.lower()
        
        # Get description from row
        description = ' '.join([str(cell) if cell else '' for cell in row[1:-1] if cell])
        description = re.sub(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', '', description)
        description = re.sub(r'[\d,]+\.\d{2}', '', description)
        description = re.sub(r'\s+', ' ', description).strip()
        
        amount = float(amounts[0].replace(',', ''))
        
        # Parse date
        try:
            if '/' in date_str:
                date_parts = date_str.split('/')
            else:
                date_parts = date_str.split('-')
            
            month = int(date_parts[0])
            day = int(date_parts[1])
            year = int(date_parts[2])
            if year < 100:
                year += 2000
            
            parsed_date = datetime(year, month, day)
            date_formatted = parsed_date.strftime('%Y-%m-%d')
        except:
            date_formatted = date_str
        
        return {
            'Date': date_formatted,
            'Type': 'Deposit' if is_deposit else 'Withdrawal',
            'Description': description,
            'Amount': amount if is_deposit else -amount,
            'Raw_Line': row_text
        }
    
    def process(self):
        """Main processing function"""
        print(f"Processing PDF: {self.pdf_path}")
        
        # Try table extraction first (more reliable for structured PDFs)
        print("Attempting table extraction...")
        transactions = self.extract_table_transactions(None)
        
        # If table extraction didn't work well, try text extraction
        if len(transactions) < 5:
            print("Table extraction yielded few results, trying text extraction...")
            text = self.extract_text_from_pdf()
            if text:
                text_transactions = self.parse_transactions(text)
                # Merge and deduplicate
                all_transactions = transactions + text_transactions
                # Remove duplicates based on date and amount
                seen = set()
                unique_transactions = []
                for txn in all_transactions:
                    key = (txn['Date'], txn['Amount'], txn['Description'][:50])
                    if key not in seen:
                        seen.add(key)
                        unique_transactions.append(txn)
                transactions = unique_transactions
        
        self.transactions = transactions
        print(f"Extracted {len(transactions)} transactions")
        
        return transactions
    
    def export_to_excel(self):
        """Export transactions to Excel file"""
        if not self.transactions:
            print("No transactions to export")
            return None
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Generate output filename
        pdf_basename = os.path.basename(self.pdf_path)
        pdf_name = os.path.splitext(pdf_basename)[0]
        output_filename = f"{pdf_name}_transactions.xlsx"
        output_path = os.path.join(self.output_dir, output_filename)
        
        # Create DataFrame
        df = pd.DataFrame(self.transactions)
        
        # Sort by date
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date')
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        
        # Reorder columns
        column_order = ['Date', 'Type', 'Description', 'Amount']
        if 'Raw_Line' in df.columns:
            column_order.append('Raw_Line')
        
        df = df[[col for col in column_order if col in df.columns]]
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Transactions']
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format amount column (make negative amounts red)
            amount_col = None
            for idx, col in enumerate(column_order, 1):
                if col == 'Amount':
                    amount_col = idx
                    break
            
            if amount_col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=amount_col)
                    if cell.value and cell.value < 0:
                        cell.font = Font(color="FF0000")
                    cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for idx, col in enumerate(column_order, 1):
                if col in df.columns:
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)
        
        print(f"Excel file created: {output_path}")
        return output_path


def _sanitize_sheet_name(name, max_len=31):
    """Create valid Excel sheet name (max 31 chars, no \\ / * ? : [ ])."""
    invalid = set('\\/*?:[]')
    clean = ''.join(c if c not in invalid else '_' for c in str(name))
    base = os.path.splitext(clean)[0].strip() or 'Sheet'
    return base[:max_len]


def _write_formatted_sheet(worksheet, df, column_order, sheet_title=''):
    """Apply header/style/width formatting to a transactions sheet."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    amount_col = None
    for idx, col in enumerate(column_order, 1):
        if col == 'Amount':
            amount_col = idx
            break
    if amount_col:
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=amount_col)
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(color="FF0000")
            cell.number_format = '#,##0.00'
    for idx, col in enumerate(column_order, 1):
        if col in df.columns:
            max_length = max(df[col].astype(str).apply(len).max(), len(col))
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)


def create_combine_all_xls(output_dir, statement_data_list):
    """
    Create combine_all_xls.xlsx with one tab per statement and a combine_data tab.
    statement_data_list: list of dicts with keys 'pdf_name', 'transactions'.
    """
    if not statement_data_list:
        return None
    
    combine_path = os.path.join(output_dir, "combine_all_xls.xlsx")
    wb = Workbook()
    # Remove default sheet; we'll add our own.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    column_order = ['Date', 'Type', 'Description', 'Amount']
    all_combined = []
    used_sheet_names = set()
    
    for item in statement_data_list:
        pdf_name = item['pdf_name']
        transactions = item['transactions']
        if not transactions:
            continue
        
        base = os.path.splitext(pdf_name)[0]
        sheet_name = _sanitize_sheet_name(base)
        orig = sheet_name
        idx = 0
        while sheet_name in used_sheet_names:
            idx += 1
            suffix = f"_{idx}"
            sheet_name = (orig[: 31 - len(suffix)] + suffix) if len(orig) + len(suffix) > 31 else orig + suffix
        used_sheet_names.add(sheet_name)
        
        df = pd.DataFrame(transactions)
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date')
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        cols = [c for c in column_order if c in df.columns]
        if 'Raw_Line' in df.columns:
            cols.append('Raw_Line')
        df = df[[c for c in cols if c in df.columns]]
        
        ws = wb.create_sheet(title=sheet_name)
        for c, col in enumerate(cols, 1):
            ws.cell(row=1, column=c, value=col)
        for r_idx, tup in enumerate(df.itertuples(index=False), 2):
            for c_idx, col in enumerate(cols, 1):
                val = getattr(tup, col, None)
                ws.cell(row=r_idx, column=c_idx, value=val)
        _write_formatted_sheet(ws, df, cols, sheet_name)
        
        for t in transactions:
            rec = dict(t)
            rec['Source'] = base
            all_combined.append(rec)
    
    if not all_combined:
        wb.save(combine_path)
        print(f"Combined workbook saved: {combine_path} (statement tabs only, no transactions)")
        return combine_path
    
    # combine_data tab
    cdf = pd.DataFrame(all_combined)
    combine_cols = ['Date', 'Type', 'Description', 'Amount', 'Source']
    if 'Raw_Line' in cdf.columns:
        combine_cols.append('Raw_Line')
    cdf = cdf[[c for c in combine_cols if c in cdf.columns]]
    cdf['Date'] = pd.to_datetime(cdf['Date'], errors='coerce')
    cdf = cdf.sort_values('Date')
    cdf['Date'] = cdf['Date'].dt.strftime('%Y-%m-%d')
    
    ws_combine = wb.create_sheet(title='combine_data', index=0)
    for c, col in enumerate(combine_cols, 1):
        ws_combine.cell(row=1, column=c, value=col)
    for r_idx, tup in enumerate(cdf.itertuples(index=False), 2):
        for c_idx, col in enumerate(combine_cols, 1):
            val = getattr(tup, col, None)
            ws_combine.cell(row=r_idx, column=c_idx, value=val)
    _write_formatted_sheet(ws_combine, cdf, combine_cols, 'combine_data')
    
    wb.save(combine_path)
    print(f"Combined workbook saved: {combine_path}")
    return combine_path


def main():
    """Main entry point - processes all PDF files in the input directory"""
    input_dir = "/Volumes/D/Praveen_Mac_D/InfiTrends/InfiTrends_Bank_Stmts"
    output_dir = "/Volumes/D/Praveen_Mac_D/InfiTrends/InfiTrends_Bank_Stmts/Stmt_Excels"
    
    # Check if input directory exists
    if not os.path.exists(input_dir):
        print(f"Error: Input directory not found at {input_dir}")
        return
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Find all PDF files in the input directory
    pdf_pattern = os.path.join(input_dir, "*.pdf")
    pdf_files = glob.glob(pdf_pattern)
    
    if not pdf_files:
        print(f"No PDF files found in {input_dir}")
        return
    
    print(f"Found {len(pdf_files)} PDF file(s) to process\n")
    print("=" * 80)
    
    # Process each PDF file -> one Excel file per statement
    successful = 0
    failed = 0
    statement_data_list = []
    
    for i, pdf_path in enumerate(pdf_files, 1):
        pdf_name = os.path.basename(pdf_path)
        print(f"\n[{i}/{len(pdf_files)}] Processing: {pdf_name}")
        print("-" * 80)
        
        try:
            parser = BankStatementParser(pdf_path, output_dir)
            transactions = parser.process()
            
            if transactions:
                output_path = parser.export_to_excel()
                print(f"✓ Success! Extracted {len(transactions)} transactions")
                print(f"  Output saved to: {output_path}")
                successful += 1
                statement_data_list.append({'pdf_name': pdf_name, 'transactions': transactions})
            else:
                print(f"⚠ Warning: No transactions were extracted from {pdf_name}")
                failed += 1
                
        except Exception as e:
            print(f"✗ Error processing {pdf_name}: {str(e)}")
            failed += 1
    
    # Build combine_all_xls: one tab per statement + combine_data
    if statement_data_list:
        print("\n" + "-" * 80)
        print("Building combine_all_xls.xlsx (all statement tabs + combine_data)...")
        create_combine_all_xls(output_dir, statement_data_list)
    
    # Summary
    print("\n" + "=" * 80)
    print(f"\nProcessing complete!")
    print(f"  Successful: {successful}")
    print(f"  Failed: {failed}")
    print(f"  Total: {len(pdf_files)}")
    print(f"\nOutput directory: {output_dir}")


if __name__ == "__main__":
    main()
