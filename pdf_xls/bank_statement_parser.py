#!/usr/bin/env python3
"""
Bank of America Statement Parser
Extracts transactions from PDF statements and exports to Excel
"""

import pdfplumber
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd


class BankStatementParser:
    def __init__(self, pdf_path, output_dir):
        self.pdf_path = pdf_path
        self.output_dir = output_dir
        self.transactions = []
        
    def extract_text_from_pdf(self):
        """Extract all text from the PDF file"""
        text_content = []
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_content.append(text)
            return '\n'.join(text_content)
        except Exception as e:
            print(f"Error reading PDF: {e}")
            return None
    
    def parse_transactions(self, text):
        """Parse transactions from extracted text"""
        transactions = []
        
        # Patterns to identify transaction sections
        deposit_pattern = r'Deposits and other credits'
        withdrawal_pattern = r'Withdrawals and other debits'
        withdrawal_continued_pattern = r'Withdrawals and other debits\s*-\s*continued'
        
        lines = text.split('\n')
        
        # Find section markers
        in_deposits = False
        in_withdrawals = False
        in_withdrawals_continued = False
        
        for i, line in enumerate(lines):
            # Check for section headers
            if re.search(deposit_pattern, line, re.IGNORECASE):
                in_deposits = True
                in_withdrawals = False
                in_withdrawals_continued = False
                continue
            elif re.search(withdrawal_continued_pattern, line, re.IGNORECASE):
                in_withdrawals_continued = True
                in_deposits = False
                in_withdrawals = False
                continue
            elif re.search(withdrawal_pattern, line, re.IGNORECASE):
                in_withdrawals = True
                in_deposits = False
                in_withdrawals_continued = False
                continue
            
            # Check if we've moved to a new section (like Summary, Balance, etc.)
            if re.search(r'^\s*(Summary|Balance|Account|Statement|Page)', line, re.IGNORECASE):
                if not any([in_deposits, in_withdrawals, in_withdrawals_continued]):
                    continue
                # Check if this is actually a transaction line or a section header
                if not self._looks_like_transaction(line):
                    in_deposits = False
                    in_withdrawals = False
                    in_withdrawals_continued = False
                    continue
            
            # Parse transaction lines
            if in_deposits or in_withdrawals or in_withdrawals_continued:
                transaction = self._parse_transaction_line(line, in_deposits)
                if transaction:
                    transactions.append(transaction)
        
        return transactions
    
    def _looks_like_transaction(self, line):
        """Check if a line looks like a transaction"""
        # Transactions typically have dates and amounts
        date_pattern = r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}'
        amount_pattern = r'[\d,]+\.\d{2}'
        
        has_date = bool(re.search(date_pattern, line))
        has_amount = bool(re.search(amount_pattern, line))
        
        return has_date and has_amount
    
    def _parse_transaction_line(self, line, is_deposit):
        """Parse a single transaction line"""
        # Skip empty lines or lines that don't look like transactions
        if not line.strip() or len(line.strip()) < 10:
            return None
        
        # Pattern for date (MM/DD/YYYY or MM/DD/YY)
        date_pattern = r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'
        # Pattern for amount (with commas and decimals)
        amount_pattern = r'([\d,]+\.\d{2})'
        
        # Find date
        date_match = re.search(date_pattern, line)
        if not date_match:
            return None
        
        date_str = date_match.group(1)
        
        # Find amount (usually at the end of the line)
        amount_matches = re.findall(amount_pattern, line)
        if not amount_matches:
            return None
        
        # The last number is usually the balance, second to last is the transaction amount
        # Or the first amount after the date might be the transaction amount
        amount_str = amount_matches[-1] if len(amount_matches) == 1 else amount_matches[-2]
        amount = float(amount_str.replace(',', ''))
        
        # Extract description (everything between date and amount)
        date_end = date_match.end()
        amount_start = line.rfind(amount_str)
        
        description = line[date_end:amount_start].strip()
        # Clean up description
        description = re.sub(r'\s+', ' ', description)
        
        # Determine transaction type
        transaction_type = 'Deposit' if is_deposit else 'Withdrawal'
        
        # Parse date
        try:
            if '/' in date_str:
                date_parts = date_str.split('/')
            else:
                date_parts = date_str.split('-')
            
            month = int(date_parts[0])
            day = int(date_parts[1])
            year = int(date_parts[2])
            if year < 100:
                year += 2000
            
            parsed_date = datetime(year, month, day)
        except:
            parsed_date = date_str
        
        return {
            'Date': parsed_date.strftime('%Y-%m-%d') if isinstance(parsed_date, datetime) else date_str,
            'Type': transaction_type,
            'Description': description,
            'Amount': amount if is_deposit else -amount,  # Negative for withdrawals
            'Raw_Line': line.strip()
        }
    
    def extract_table_transactions(self, text):
        """Alternative method: Try to extract transactions using table detection"""
        transactions = []
        
        try:
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # Try to extract tables
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if not table:
                            continue
                        
                        # Look for transaction tables
                        for row in table:
                            if not row or len(row) < 3:
                                continue
                            
                            # Check if row contains transaction data
                            row_text = ' '.join([str(cell) if cell else '' for cell in row])
                            
                            # Check for date pattern
                            if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', row_text):
                                transaction = self._parse_table_row(row, row_text)
                                if transaction:
                                    transactions.append(transaction)
        except Exception as e:
            print(f"Error extracting tables: {e}")
        
        return transactions
    
    def _parse_table_row(self, row, row_text):
        """Parse a transaction from a table row"""
        # Find date
        date_match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', row_text)
        if not date_match:
            return None
        
        date_str = date_match.group(1)
        
        # Find amounts
        amount_pattern = r'([\d,]+\.\d{2})'
        amounts = re.findall(amount_pattern, row_text)
        if not amounts:
            return None
        
        # Determine if deposit or withdrawal based on context
        is_deposit = 'deposit' in row_text.lower() or 'credit' in row_text.lower()
        
        # Get description from row
        description = ' '.join([str(cell) if cell else '' for cell in row[1:-1] if cell])
        description = re.sub(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', '', description)
        description = re.sub(r'[\d,]+\.\d{2}', '', description)
        description = re.sub(r'\s+', ' ', description).strip()
        
        amount = float(amounts[0].replace(',', ''))
        
        # Parse date
        try:
            if '/' in date_str:
                date_parts = date_str.split('/')
            else:
                date_parts = date_str.split('-')
            
            month = int(date_parts[0])
            day = int(date_parts[1])
            year = int(date_parts[2])
            if year < 100:
                year += 2000
            
            parsed_date = datetime(year, month, day)
            date_formatted = parsed_date.strftime('%Y-%m-%d')
        except:
            date_formatted = date_str
        
        return {
            'Date': date_formatted,
            'Type': 'Deposit' if is_deposit else 'Withdrawal',
            'Description': description,
            'Amount': amount if is_deposit else -amount,
            'Raw_Line': row_text
        }
    
    def process(self):
        """Main processing function"""
        print(f"Processing PDF: {self.pdf_path}")
        
        # Try table extraction first (more reliable for structured PDFs)
        print("Attempting table extraction...")
        transactions = self.extract_table_transactions(None)
        
        # If table extraction didn't work well, try text extraction
        if len(transactions) < 5:
            print("Table extraction yielded few results, trying text extraction...")
            text = self.extract_text_from_pdf()
            if text:
                text_transactions = self.parse_transactions(text)
                # Merge and deduplicate
                all_transactions = transactions + text_transactions
                # Remove duplicates based on date and amount
                seen = set()
                unique_transactions = []
                for txn in all_transactions:
                    key = (txn['Date'], txn['Amount'], txn['Description'][:50])
                    if key not in seen:
                        seen.add(key)
                        unique_transactions.append(txn)
                transactions = unique_transactions
        
        self.transactions = transactions
        print(f"Extracted {len(transactions)} transactions")
        
        return transactions
    
    def export_to_excel(self):
        """Export transactions to Excel file"""
        if not self.transactions:
            print("No transactions to export")
            return None
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Generate output filename
        pdf_basename = os.path.basename(self.pdf_path)
        pdf_name = os.path.splitext(pdf_basename)[0]
        output_filename = f"{pdf_name}_transactions.xlsx"
        output_path = os.path.join(self.output_dir, output_filename)
        
        # Create DataFrame
        df = pd.DataFrame(self.transactions)
        
        # Sort by date
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date')
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        
        # Reorder columns
        column_order = ['Date', 'Type', 'Description', 'Amount']
        if 'Raw_Line' in df.columns:
            column_order.append('Raw_Line')
        
        df = df[[col for col in column_order if col in df.columns]]
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Transactions']
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format amount column (make negative amounts red)
            amount_col = None
            for idx, col in enumerate(column_order, 1):
                if col == 'Amount':
                    amount_col = idx
                    break
            
            if amount_col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=amount_col)
                    if cell.value and cell.value < 0:
                        cell.font = Font(color="FF0000")
                    cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for idx, col in enumerate(column_order, 1):
                if col in df.columns:
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)
        
        print(f"Excel file created: {output_path}")
        return output_path


def main():
    """Main entry point"""
    pdf_path = "/Volumes/D/Praveen_Mac_D/InfiTrends/InfiTrends_Bank_Stmts/InfiTrends_eStmt_2025-12-31.pdf"
    output_dir = "/Volumes/D/Praveen_Mac_D/InfiTrends/InfiTrends_Bank_Stmts/Stmt_Excels"
    
    # Check if PDF exists
    if not os.path.exists(pdf_path):
        print(f"Error: PDF file not found at {pdf_path}")
        return
    
    # Create parser and process
    parser = BankStatementParser(pdf_path, output_dir)
    transactions = parser.process()
    
    if transactions:
        output_path = parser.export_to_excel()
        print(f"\nSuccess! Extracted {len(transactions)} transactions")
        print(f"Output saved to: {output_path}")
    else:
        print("\nWarning: No transactions were extracted. Please check the PDF format.")


if __name__ == "__main__":
    main()
